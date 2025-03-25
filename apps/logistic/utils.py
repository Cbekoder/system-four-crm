import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Activity Report"

# Set page layout to mimic the image (A4 size, portrait orientation)
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

# Define some styles
bold_font = Font(bold=True, size=12)
center_alignment = Alignment(horizontal='center', vertical='center')
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

# Merge cells and set header information
ws.merge_cells('B3:N3')
ws['B3'] = "1413"
ws['B3'].font = Font(bold=True, size=16)
ws['B3'].alignment = center_alignment

ws.merge_cells('B4:N4')
ws['B4'] = "20.03.2025 5инт пубый"
ws['B4'].alignment = center_alignment

ws.merge_cells('B5:G5')
ws['B5'] = "20.03.2025 17:00"
ws['B5'].alignment = center_alignment

ws.merge_cells('I5:N5')
ws['I5'] = "19.04.2025 24:00"
ws['I5'].alignment = center_alignment

# Personal and vehicle information
ws['B7'] = "АБДУЛЛАЕВ АЗАМАТХОЖИ"
ws['B7'].font = bold_font

ws['B8'] = "ДАФ 40 148 УБА"

ws['B9'] = "БОНУМ 40 0055 СА"

ws['B10'] = "АА№0153428"
ws['B10'].alignment = Alignment(horizontal='right')

# Company information
ws.merge_cells('B12:N12')
ws['B12'] = "ООО ОЛТАРИО КОМФОРТ ТРАНС"
ws['B12'].font = bold_font
ws['B12'].alignment = center_alignment

ws.merge_cells('B13:N13')
ws['B13'] = "page 1"
ws['B13'].alignment = center_alignment

# Footer information
ws.merge_cells('B15:N15')
ws['B15'] = "ПО ТЕРРИТОРИИ СНГ"
ws['B15'].alignment = center_alignment

ws['B16'] = "Р.СОХНБАЗАРОВ"
ws['N16'] = "М.ЖК"
ws['N16'].alignment = Alignment(horizontal='right')

# Add green dots (mimicking the positions in the image)
# These positions are approximate based on the image
green_dot_positions = [
    (3, 2),   # Near the QR code
    (6, 2),   # Left side
    (9, 2),   # Left side
    (12, 3),  # Slightly right
    (14, 4),  # Further right
    (16, 5),  # Further right
    (18, 6),  # Further right
    (20, 7),  # Further right
    (22, 8),  # Further right
    (24, 9),  # Further right
    (26, 10), # Further right
    (28, 11), # Further right
]

for row, col in green_dot_positions:
    cell = ws.cell(row=row, column=col)
    cell.fill = green_fill

# Placeholder for QR code (we'll just leave a space for it)
ws.merge_cells('B2:C3')
ws['B2'] = "[QR Code Placeholder]"
ws['B2'].alignment = center_alignment

# Adjust column widths and row heights for better layout
for col in range(1, 15):  # Columns A to N
    ws.column_dimensions[get_column_letter(col)].width = 5
for row in range(1, 30):  # Rows 1 to 29
    ws.row_dimensions[row].height = 20

# Save the file
wb.save("activity_report.xlsx")
print("Excel file 'activity_report.xlsx' has been created.")